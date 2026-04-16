"""Report generation: CSV, XLSX, and terminal summaries."""
from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path

from .models import DuplicateEntry, RunStats, SearchResult

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    HAS_XLSX = True
except Exception:
    HAS_XLSX = False

REPORT_FIELDS = [
    "song",
    "source",
    "source_used",
    "status",
    "score",
    "cache_hit",
    "cache_hits",
    "matched_query",
    "matched_query_kind",
    "fallback_used",
    "resolved_phase",
    "result_tier",
    "candidate_title",
    "best_seen_source",
    "best_seen_score",
    "best_seen_tier",
    "best_seen_url",
    "page_url",
    "direct_url",
    "saved_file",
    "note",
]


def _result_to_row(r: SearchResult) -> dict[str, str]:
    return {
        "song": r.song,
        "source": r.source,
        "source_used": r.source,
        "status": r.status.value if hasattr(r.status, "value") else str(r.status),
        "score": f"{r.score:.3f}",
        "cache_hit": "yes" if r.cache_hit else "no",
        "cache_hits": str(r.cache_hits),
        "matched_query": r.matched_query,
        "matched_query_kind": r.matched_query_kind,
        "fallback_used": "yes" if r.fallback_used else "no",
        "resolved_phase": r.resolved_phase,
        "result_tier": r.result_tier.value if hasattr(r.result_tier, "value") else str(r.result_tier),
        "candidate_title": r.candidate_title,
        "best_seen_source": r.best_seen_source,
        "best_seen_score": f"{r.best_seen_score:.3f}" if r.best_seen_score else "",
        "best_seen_tier": r.best_seen_tier,
        "best_seen_url": r.best_seen_url,
        "page_url": r.page_url,
        "direct_url": r.direct_url or "",
        "saved_file": r.saved_file,
        "note": r.note,
    }


def result_to_row(result: SearchResult) -> dict[str, str]:
    return _result_to_row(result)


def save_csv(results: list[SearchResult], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=REPORT_FIELDS)
        writer.writeheader()
        for r in results:
            writer.writerow(_result_to_row(r))


def save_xlsx(results: list[SearchResult], path: Path) -> bool:
    if not HAS_XLSX:
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "report"

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    ws.append(REPORT_FIELDS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Status color map
    status_colors: dict[str, str] = {
        "downloaded": "C6EFCE",
        "page_found": "FFEB9C",
        "blocked": "FFC7CE",
        "not_found": "F2F2F2",
        "download_error": "FFC7CE",
        "error": "FFC7CE",
    }

    for r in results:
        row = _result_to_row(r)
        ws.append([row.get(f, "") for f in REPORT_FIELDS])
        color = status_colors.get(row["status"], "FFFFFF")
        fill = PatternFill("solid", fgColor=color)
        for cell in ws[ws.max_row]:
            cell.fill = fill

    # Auto-width (approximate)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    wb.save(path)
    return True


def save_duplicates_csv(duplicates: list[DuplicateEntry], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["raw_song", "matched_song", "reason"])
        for d in duplicates:
            writer.writerow([d.raw_song, d.matched_song, d.reason])


def save_errors_log(errors: list[str], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for e in errors:
            f.write(f"[{stamp}] {e}\n")


def save_summary_json(summary: dict[str, object], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")


def format_elapsed(seconds: float) -> str:
    total_seconds = max(0, int(round(seconds)))
    hours, remainder = divmod(total_seconds, 3600)
    minutes, secs = divmod(remainder, 60)
    if hours:
        return f"{hours}h {minutes:02d}m {secs:02d}s"
    if minutes:
        return f"{minutes}m {secs:02d}s"
    return f"{secs}s"


def print_summary(stats: RunStats, paths: dict[str, Path | None], use_color: bool = True) -> None:
    from .logging_utils import Printer

    p = Printer(color=use_color)
    p.separator()
    p.bold("SUMMARY")

    useful = stats.downloaded + stats.page_found
    useful_pct = f"  ({useful * 100 // max(stats.total, 1)}%)" if stats.total else ""

    p.info(f"total processed        : {stats.total}")
    if stats.duplicates:
        p.info(f"duplicates skipped     : {stats.duplicates}")
    p.ok(  f"downloaded             : {stats.downloaded}")
    p.warn(f"page_found             : {stats.page_found}")
    p.info(f"useful results         : {useful}{useful_pct}")
    p.info(f"not_found              : {stats.not_found}")
    if stats.blocked:
        p.warn(f"blocked (403/429)      : {stats.blocked}")
    if stats.download_error:
        p.err( f"download errors        : {stats.download_error}")
    if stats.errors:
        p.err( f"errors                 : {stats.errors}")

    if stats.elapsed_seconds:
        p.info(f"elapsed                : {format_elapsed(stats.elapsed_seconds)}")
    if stats.avg_seconds_per_song:
        p.info(f"avg per song           : {stats.avg_seconds_per_song:.1f}s")
    if stats.avg_seconds_per_success and useful:
        p.info(f"avg per useful result  : {stats.avg_seconds_per_success:.1f}s")

    p.info(f"phase wins             : A={stats.phase_a_wins}, B={stats.phase_b_wins}")

    p.separator()
    for label, path in paths.items():
        if path and path.exists():
            p.info(f"{label:<22} : {path}")
